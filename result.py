from openpyxl import workbook, load_workbook
import PyPDF2
import pathlib

print("Ievadiet: Rēķins_Nr_1.pdf, Rēķins_Nr_2.pdf vai Rēķins_Nr_3.pdf ")
data=[]
file= input()
adrese = pathlib.Path(file)
tarifs=0
menesis=0
r=[]
saraksts=[]
row=[]


if file==("Rēķins_Nr_1.pdf") or file==("Rēķins_Nr_2.pdf") or file==("Rēķins_Nr_3.pdf"):
    pdf_file=PyPDF2.PdfReader(open(adrese,"rb"))
    number_of_pages=len(pdf_file.pages)
    page1=pdf_file.pages[0]
    

    text1=page1.extract_text()
    


    
    pos1 = text1.find("Kopā apmaksai")
    
    summa = text1[pos1+14:pos1+19].rstrip()
    summa = float(summa.replace(',', '.'))
    row.append(summa)
    

    pos1 = text1.find("Datums: ")
    datums = text1[pos1+11:pos1+13]
    datums=datums.rstrip().replace(" ","")
    row.append(datums)
    

    with open("menesis.csv","r") as f:
        next(f)
        for line in f:
            r=line.rstrip().split(",")
            x=r[1].rstrip().replace(" ","")
            

            if datums == r[1]:
                menesis=r[0]
                saraksts.append(menesis)

    excel_files = ["Report_01.09.2023_30.09.2023.xlsx", "Report_01.08.2023_31.08.2023.xlsx", "Report_01.07.2023_31.07.2023.xlsx"]

    for excel_file in excel_files:
        wb = load_workbook(excel_file)
        ws = wb.active
        max_row = ws.max_row
        excel_menesis= ws['F' + str(2)].value
        saraksts.append(excel_menesis)

        if menesis==excel_menesis:
            apmeklejums=[]
            max_row = ws.max_row

            for row in range(3,max_row+1):
                datumi= ws['C' + str(row)].value
                apmeklejums.append(datumi)
                apmeklejums_skaits=len(apmeklejums)
            
            maksa_par_apmeklejumu=round((summa/apmeklejums_skaits), 2)
            print(f"Sporta klubā 'MyFitness' maksa par vienu apmeklējumu ir {maksa_par_apmeklejumu} EUR, kamēr RTU sporta zālē Meža ielā- 3 EUR.")
            if maksa_par_apmeklejumu< 3:
                print("Ir izdevīgi saglabāt esošo MyFitness abonementu.")
            else:
                print("Nav izdevīgi saglabāt esošo MyFitness abonementu.")
    wb.close()